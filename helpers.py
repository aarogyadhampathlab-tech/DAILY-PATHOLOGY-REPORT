import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import config

OPENAI_AVAILABLE = config.OPENAI_AVAILABLE
openai = getattr(config, 'openai', None)
CATEGORY_RULES = config.CATEGORY_RULES


def normalize_bookingmode(x):
    s = "" if pd.isna(x) else str(x).strip().upper()
    if "IPD" in s:
        return "IPD"
    return "OPD Indent"


def ai_batch_categorize(unknown_tests):
    if not unknown_tests or not OPENAI_AVAILABLE or not openai or not getattr(openai, 'api_key', None):
        return {}
    tests_text = "\n".join([f"- Test: {t}, Subgroup: {s}" for t, s in unknown_tests])
    prompt = f"Categories: Biochemistry, Clinical, Hematology, Immunology. Assign each test. Return CSV: TestName,Subgroup,Category\n{tests_text}"
    try:
        resp = openai.ChatCompletion.create(model="gpt-4o-mini", messages=[{"role": "user", "content": prompt}], temperature=0)
        mapping = {}
        for line in resp['choices'][0]['message']['content'].strip().splitlines():
            parts = [p.strip() for p in line.split(",")]
            if len(parts) == 3:
                mapping[(parts[0], parts[1])] = parts[2]
        return mapping
    except Exception:
        return {}


def build_test_counts(df):
    df = df.copy()
    if 'BookingMode' in df.columns:
        df["BookingMode_norm"] = df["BookingMode"].apply(normalize_bookingmode)
    else:
        df["BookingMode_norm"] = "OPD Indent"
    pivot = df.pivot_table(index="TestName", columns="BookingMode_norm", aggfunc="size", fill_value=0).reset_index()
    pivot["IPD"] = pivot.get("IPD", 0)
    pivot["OPD"] = pivot.get("OPD Indent", 0)
    pivot["Total"] = pivot[["IPD", "OPD"]].sum(axis=1)
    result = pivot[["TestName", "IPD", "OPD", "Total"]].sort_values("TestName").reset_index(drop=True)
    grand_total = pd.DataFrame([{"TestName": "Grand Total", "IPD": int(result["IPD"].sum()), "OPD": int(result["OPD"].sum()), "Total": int(result["Total"].sum())}])
    return pd.concat([result, grand_total], ignore_index=True)


def build_category_counts(df):
    df = df.copy()
    unknown_tests = []
    final_cats = []
    for _, row in df.iterrows():
        text = f"{row['TestName']} {row['subgroup']}".upper()
        final = next((cat for cat, keys in CATEGORY_RULES.items() if any(k.upper() in text for k in keys)), None)
        if not final:
            unknown_tests.append((str(row['TestName']), str(row['subgroup'])))
        final_cats.append(final)
    ai_mapping = ai_batch_categorize(unknown_tests)
    df["Final_Category"] = [c or ai_mapping.get((str(r.TestName), str(r.subgroup)), "Biochemistry") for c, r in zip(final_cats, df.itertuples())]
    results = [{"Category": c, "Count": int((df["Final_Category"] == c).sum())} for c in CATEGORY_RULES.keys()]
    results.append({"Category": "Grand Total", "Count": int(len(df))})
    return pd.DataFrame(results), df


def style_excel(test_counts, cat_counts, report_date_str):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        test_counts.to_excel(writer, sheet_name="Analysis", startrow=2, index=False)
        start_cat = len(test_counts) + 5
        cat_counts.to_excel(writer, sheet_name="Analysis", startrow=start_cat+1, index=False)

    wb = load_workbook(filename=BytesIO(output.getvalue()))
    ws = wb.active

    thin_side = Side(border_style="thin", color="000000")
    full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    total_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"DAILY PATHOLOGY REPORT - {report_date_str}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=3, max_row=3 + len(test_counts), min_col=1, max_col=4):
        for cell in row:
            cell.border = full_border
            if cell.row == 3:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if "Grand Total" in str(ws.cell(row=cell.row, column=1).value):
                cell.fill = total_fill
                cell.font = Font(bold=True)

    cat_header_idx = start_cat + 2
    for row in ws.iter_rows(min_row=cat_header_idx, max_row=cat_header_idx + len(cat_counts), min_col=1, max_col=2):
        for cell in row:
            cell.border = full_border
            if cell.row == cat_header_idx:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if "Grand Total" in str(ws.cell(row=cell.row, column=1).value):
                cell.fill = total_fill
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 12

    final_output = BytesIO()
    wb.save(final_output)
    return final_output

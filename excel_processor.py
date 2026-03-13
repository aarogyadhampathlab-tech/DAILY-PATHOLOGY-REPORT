import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import os
import pickle  # For saving/loading DataFrames
# ------------------------- (existing code remains unchanged up to here)

# ------------------------- (add new section for data persistence)
DATA_DIR = "processed_data"
os.makedirs(DATA_DIR, exist_ok=True)

def save_processed_data(date_str, df, test_counts, cat_counts):
    """Save processed data for a given date."""
    date_dir = os.path.join(DATA_DIR, date_str)
    os.makedirs(date_dir, exist_ok=True)
    df.to_csv(os.path.join(date_dir, "raw.csv"), index=False)
    test_counts.to_csv(os.path.join(date_dir, "test_counts.csv"), index=False)
    cat_counts.to_csv(os.path.join(date_dir, "cat_counts.csv"), index=False)

def load_processed_data(date_str):
    """Load processed data for a given date."""
    date_dir = os.path.join(DATA_DIR, date_str)
    if not os.path.exists(date_dir):
        return None, None, None
    df = pd.read_csv(os.path.join(date_dir, "raw.csv"))
    test_counts = pd.read_csv(os.path.join(date_dir, "test_counts.csv"))
    cat_counts = pd.read_csv(os.path.join(date_dir, "cat_counts.csv"))
    return df, test_counts, cat_counts

def delete_processed_data(date_str):
    """Delete processed data for a given date."""
    date_dir = os.path.join(DATA_DIR, date_str)
    if os.path.exists(date_dir):
        import shutil
        shutil.rmtree(date_dir)

def get_saved_dates():
    """Get list of saved dates."""
    if not os.path.exists(DATA_DIR):
        return []
    return sorted([d for d in os.listdir(DATA_DIR) if os.path.isdir(os.path.join(DATA_DIR, d))])

def compute_cumulative():
    """Compute cumulative sums across all saved dates."""
    dates = get_saved_dates()
    if not dates:
        return None, None
    all_test_counts = []
    all_cat_counts = []
    for date in dates:
        _, tc, cc = load_processed_data(date)
        if tc is not None:
            tc['Date'] = date
            all_test_counts.append(tc)
        if cc is not None:
            cc['Date'] = date
            all_cat_counts.append(cc)
    if all_test_counts:
        combined_tc = pd.concat(all_test_counts, ignore_index=True)
        cumulative_tc = combined_tc.groupby('TestName')[['IPD', 'OPD', 'Total']].sum().reset_index()
        grand_total_tc = pd.DataFrame([{"TestName": "Grand Total", "IPD": int(cumulative_tc["IPD"].sum()), "OPD": int(cumulative_tc["OPD"].sum()), "Total": int(cumulative_tc["Total"].sum())}])
        cumulative_tc = pd.concat([cumulative_tc, grand_total_tc], ignore_index=True)
    else:
        cumulative_tc = None
    if all_cat_counts:
        combined_cc = pd.concat(all_cat_counts, ignore_index=True)
        cumulative_cc = combined_cc.groupby('Category')[['Count']].sum().reset_index()
        grand_total_cc = pd.DataFrame([{"Category": "Grand Total", "Count": int(cumulative_cc["Count"].sum())}])
        cumulative_cc = pd.concat([cumulative_cc, grand_total_cc], ignore_index=True)
    else:
        cumulative_cc = None
    return cumulative_tc, cumulative_cc

# ------------------------- (modify the execution section)
# ------------------------- (existing code up to "if uploaded_file:" remains)

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    test_counts = build_test_counts(df)
    cat_counts, categorized_df = build_category_counts(df)
    
    # Save the processed data
    save_processed_data(yesterday_str, df, test_counts, cat_counts)
    st.success(f"Data for {yesterday_str} saved successfully.")
    
    # Dashboard Metrics (existing)
    m1, m2, m3 = st.columns(3)
    m1.metric("Total Tests", len(df))
    m2.metric("IPD", test_counts.iloc[-1]["IPD"])
    m3.metric("OPD", test_counts.iloc[-1]["OPD"])

    # Download Button (existing)
    excel_report = style_excel(test_counts, cat_counts)
    st.sidebar.download_button(
        label="📥 Download Excel Report",
        data=excel_report.getvalue(),
        file_name=f"Pathology_Report_{yesterday_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Existing display sections (existing)
    if show_category_table:
        st.subheader("Category Summary")
        st.table(cat_counts)
    if show_test_table:
        st.subheader("Detailed Test Counts")
        st.dataframe(test_counts, use_container_width=True, hide_index=True)
    if show_raw:
        st.subheader("Raw Data Preview")
        st.dataframe(df.head(100), use_container_width=True)

# ------------------------- (add new section for managing saved data)
st.sidebar.divider()
st.sidebar.subheader("📂 Saved Data Management")

saved_dates = get_saved_dates()
if saved_dates:
    selected_date = st.sidebar.selectbox("Select Date to View/Delete", saved_dates)
    col1, col2 = st.sidebar.columns(2)
    if col1.button("View Data"):
        df_view, tc_view, cc_view = load_processed_data(selected_date)
        if df_view is not None:
            st.subheader(f"Data for {selected_date}")
            st.dataframe(tc_view, use_container_width=True, hide_index=True)
            st.table(cc_view)
        else:
            st.error("Data not found.")
    if col2.button("Delete Data"):
        delete_processed_data(selected_date)
        st.success(f"Data for {selected_date} deleted.")
        st.rerun()  # Refresh to update list
else:
    st.sidebar.write("No saved data yet.")

# ------------------------- (add cumulative visualization section)
st.divider()
st.header("📊 Cumulative Summary")
cum_tc, cum_cc = compute_cumulative()
if cum_tc is not None and cum_cc is not None:
    st.subheader("Cumulative Test Counts")
    st.dataframe(cum_tc, use_container_width=True, hide_index=True)
    st.bar_chart(cum_tc.set_index('TestName')[['IPD', 'OPD', 'Total']])
    
    st.subheader("Cumulative Category Counts")
    st.table(cum_cc)
    st.bar_chart(cum_cc.set_index('Category')['Count'])
else:
    st.info("No cumulative data available.")
# -------------------------
# Configuration & Dates
# -------------------------
# Calculate yesterday's date for report labeling
today = datetime.today()
yesterday = today - timedelta(days=1)
today_str = today.strftime('%d-%m-%Y')
yesterday_str = yesterday.strftime('%d-%m-%Y')

# Optional: AI categorization (safe fallback if unavailable)
try:
    import openai
    OPENAI_AVAILABLE = True
    openai.api_key = os.getenv("OPENAI_API_KEY")
except Exception:
    OPENAI_AVAILABLE = False

st.set_page_config(page_title="Daily Pathology Report", page_icon="📊", layout="wide")

# Custom CSS for modern look
st.markdown("""
<style>
.main { background-color: #f8f9fa; }
.block-container { padding-top: 1rem; }
[data-testid="stMetric"] {
    background-color: white;
    padding: 15px;
    border-radius: 10px;
    box-shadow: 0 0 6px rgba(0,0,0,0.08);
}
</style>
""", unsafe_allow_html=True)

# Header with dynamic date
st.markdown(f"""
<h2 style="margin-bottom:0">Daily Pathology Report</h2>
<p style="color:gray;margin-top:0">
Pathology Department · Aarogyadham Hospital <br>
<b>Report Date: {yesterday_str}</b> | Generated: {today_str}
</p>
<hr>
""", unsafe_allow_html=True)

# -------------------------
# Static category rules
# -------------------------
CATEGORY_RULES = {
    "Biochemistry": [
        "RENAL FUNCTION TEST", "LIVER FUNCTION TEST", "BLOOD GLUCOSE",
        "GLYCOSYLATED HB", "SGOT", "SGPT", "BLOOD UREA",
        "VIRAL MARKER", "PREOPERATIVE PROFILE", "SEROLOGY", "PT/INR"
    ],
    "Clinical": [
        "URINE ANALYSIS", "PLEURAL FLUID EXAMINATION",
        "Plural Fluid for R/E Biochemistry / ADA"
    ],
    "Hematology": [
        "COMPLETE BLOOD COUNTS [CBC]", "TOTAL LEUCOCYTE COUNT",
        "FLUID DLC", "COMPLETE HEMOGRAM WITH ESR", "BLOOD GROUP"
    ],
    "Immunology": [
        "Hormone Assays Report", "Serum IGE", "VDRL TITER", "HBsAg",
        "HCV ANTIBODY TEST", "CA-125", "THYROID FUNCTION TEST",
        "THYROID STIMULATING HORMONE", "TOTAL THYROID PROFILE",
        "IgG IgM S Typhe", "C-REACTIVE PROTEIN"
    ]
}

# -------------------------
# Helpers
# -------------------------
def normalize_bookingmode(x):
    s = "" if pd.isna(x) else str(x).strip().upper()
    if "IPD" in s: return "IPD"
    return "OPD Indent"

def ai_batch_categorize(unknown_tests):
    if not unknown_tests or not OPENAI_AVAILABLE or not openai.api_key:
        return {}
    tests_text = "\n".join([f"- Test: {t}, Subgroup: {s}" for t, s in unknown_tests])
    prompt = f"Categories: Biochemistry, Clinical, Hematology, Immunology. Assign each test. Return CSV: TestName,Subgroup,Category\n{tests_text}"
    try:
        resp = openai.ChatCompletion.create(model="gpt-4o-mini", messages=[{"role": "user", "content": prompt}], temperature=0)
        mapping = {}
        for line in resp['choices'][0]['message']['content'].strip().splitlines():
            parts = [p.strip() for p in line.split(",")]
            if len(parts) == 3: mapping[(parts[0], parts[1])] = parts[2]
        return mapping
    except: return {}

def build_test_counts(df):
    df = df.copy()
    df["BookingMode_norm"] = df["BookingMode"].apply(normalize_bookingmode)
    pivot = df.pivot_table(index="TestName", columns="BookingMode_norm", aggfunc="size", fill_value=0).reset_index()
    pivot["IPD"] = pivot.get("IPD", 0)
    pivot["OPD"] = pivot.get("OPD Indent", 0)
    pivot["Total"] = pivot[["IPD", "OPD"]].sum(axis=1)
    result = pivot[["TestName", "IPD", "OPD", "Total"]].sort_values("TestName").reset_index(drop=True)
    grand_total = pd.DataFrame([{"TestName": "Grand Total", "IPD": int(result["IPD"].sum()), "OPD": int(result["OPD"].sum()), "Total": int(result["Total"].sum())}])
    return pd.concat([result, grand_total], ignore_index=True)

def build_category_counts(df):
    df = df.copy()
    unknown_tests, final_cats = [], []
    for _, row in df.iterrows():
        text = f"{row['TestName']} {row['subgroup']}".upper()
        final = next((cat for cat, keys in CATEGORY_RULES.items() if any(k.upper() in text for k in keys)), None)
        if not final: unknown_tests.append((str(row['TestName']), str(row['subgroup'])))
        final_cats.append(final)
    ai_mapping = ai_batch_categorize(unknown_tests)
    df["Final_Category"] = [c or ai_mapping.get((str(r.TestName), str(r.subgroup)), "Biochemistry") for c, r in zip(final_cats, df.itertuples())]
    results = [{"Category": c, "Count": int((df["Final_Category"] == c).sum())} for c in CATEGORY_RULES.keys()]
    results.append({"Category": "Grand Total", "Count": int(len(df))})
    return pd.DataFrame(results), df

def style_excel(test_counts, cat_counts):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        test_counts.to_excel(writer, sheet_name="Analysis", startrow=2, index=False)
        start_cat = len(test_counts) + 5
        cat_counts.to_excel(writer, sheet_name="Analysis", startrow=start_cat+1, index=False)

    wb = load_workbook(filename=BytesIO(output.getvalue()))
    ws = wb.active

    # Style definitions
    thin_side = Side(border_style="thin", color="000000")
    full_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    total_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Title & Previous Day Date
    ws.merge_cells("A1:D1")
    ws["A1"] = f"DAILY PATHOLOGY REPORT - {yesterday_str}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 1. Apply Borders to Test Counts Table
    for row in ws.iter_rows(min_row=3, max_row=3 + len(test_counts), min_col=1, max_col=4):
        for cell in row:
            cell.border = full_border
            if cell.row == 3: # Header
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if "Grand Total" in str(ws.cell(row=cell.row, column=1).value):
                cell.fill = total_fill
                cell.font = Font(bold=True)

    # 2. Apply Borders to Category Table
    cat_header_idx = start_cat + 2
    for row in ws.iter_rows(min_row=cat_header_idx, max_row=cat_header_idx + len(cat_counts), min_col=1, max_col=2):
        for cell in row:
            cell.border = full_border
            if cell.row == cat_header_idx: # Header
                cell.fill = header_fill
                cell.font = Font(bold=True)
            if "Grand Total" in str(ws.cell(row=cell.row, column=1).value):
                cell.fill = total_fill
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    for col in ["B","C","D"]: ws.column_dimensions[col].width = 12

    final_output = BytesIO()
    wb.save(final_output)
    return final_output

# -------------------------
# Sidebar & Execution
# -------------------------
st.sidebar.title("🧪 Report Controls")
st.sidebar.write("Aarogyadham Hospital")
uploaded_file = st.sidebar.file_uploader("Upload Daily Excel File", type=["xlsx"])
st.sidebar.divider()
show_raw = st.sidebar.checkbox("Show Raw Data", value=False)
show_test_table = st.sidebar.checkbox("Show Test-wise Table", value=True)
show_category_table = st.sidebar.checkbox("Show Category Summary", value=True)

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    test_counts = build_test_counts(df)
    cat_counts, categorized_df = build_category_counts(df)
    
    # Dashboard Metrics
    m1, m2, m3 = st.columns(3)
    m1.metric("Total Tests", len(df))
    m2.metric("IPD", test_counts.iloc[-1]["IPD"])
    m3.metric("OPD", test_counts.iloc[-1]["OPD"])

    # Download Button
    excel_report = style_excel(test_counts, cat_counts)
    st.sidebar.download_button(
        label="📥 Download Excel Report",
        data=excel_report.getvalue(),
        file_name=f"Pathology_Report_{yesterday_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if show_category_table:
        st.subheader("Category Summary")
        st.table(cat_counts)
    if show_test_table:
        st.subheader("Detailed Test Counts")
        st.dataframe(test_counts, use_container_width=True, hide_index=True)
    if show_raw:
        st.subheader("Raw Data Preview")
        st.dataframe(df.head(100), use_container_width=True)
else:
    st.info("Upload the Daily Excel file in the sidebar to begin.")
